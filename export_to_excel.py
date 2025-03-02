import openpyxl
from openpyxl.styles import PatternFill
from PySide6.QtWidgets import QFileDialog, QInputDialog

def export_transactions_to_excel(transactions, summary):
    if not transactions:
        print("No transactions to export.")
        return

    file_path, _ = QFileDialog.getOpenFileName(None, "Select Excel File", "", "Excel Files (*.xlsx)")
    if not file_path:
        return

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"Error opening file: {e}")
        return

    sheet_names = wb.sheetnames
    sheet_name, ok = QInputDialog.getItem(None, "Select Sheet", "Choose a sheet:", sheet_names, 0, False)
    if not ok or not sheet_name:
        return

    ws = wb[sheet_name]

    existing_tickets = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            existing_tickets.add(str(row[1]))

    row_to_write = ws.max_row + 1
    for row in range(2, ws.max_row + 1):
        if not ws.cell(row=row, column=2).value:
            row_to_write = row
            break

    # Color fills
    yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for trans in transactions:
        if trans["ticket"] in existing_tickets:
            continue

        row = [
            trans["date"],       # A - Transaction Date (DD.MM.YYYY)
            trans["ticket"],     # B - Ticket
            trans["position_id"],# C - Position ID
            trans["action"],     # D - Opening/Closing
            trans["instrument"], # E - Instrument
            trans["entry_time"], # F - Entry Time (HH:MM:SS)
            "",                  # G (empty)
            trans["side"],       # H - Buy/Sell
            trans["entry_price"],# I - Entry Price
            trans["close_price"],# J - Close Price
            trans["sl_price"],   # K - SL Price
            trans["sl_pips"],    # L - SL in Pips
            trans["tp_price"],   # M - TP Price
            trans["tp_pips"],    # N - TP in Pips
            trans["volume"],     # O - Volume
            trans["result_pips"],# P - Result in Pips
            trans["result_usd"], # Q - Result in USD
            trans["category"]    # R - Win/Loss/Zero Classification
        ]

        ws.append(row)
        current_row = ws.max_row

        # Apply conditional formatting
        if trans["action"] == "Open":
            for col in range(1, len(row) + 1):
                ws.cell(row=current_row, column=col).fill = yellow_fill
        elif trans["action"] == "Close":
            fill = green_fill if float(trans["result_usd"]) > 0 else red_fill
            for col in range(1, len(row) + 1):
                ws.cell(row=current_row, column=col).fill = fill

    # Insert summary row
    summary_row = ["Summary", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""]
    ws.append(summary_row)

    summary_data = [
        ["Total Result (USD)", summary["total_result"]],
        ["Total Deals", summary["total_count"]],
        ["Win Deals", summary["win_count"], f"{(summary["win_count"] / summary["total_count"] * 100):.2f}%"],
        ["Lose Deals", summary["lose_count"], f"{(summary["lose_count"] / summary["total_count"] * 100):.2f}%"],
        ["Zero Deals", summary["zero_count"], f"{(summary["zero_count"] / summary["total_count"] * 100):.2f}%"],
        ["Win Result (USD)", summary["win_result"]],
        ["Lose Result (USD)", summary["lose_result"]],
        ["Zero Result (USD)", summary["zero_result"]]
    ]

    for row_data in summary_data:
        ws.append(row_data)

    try:
        wb.save(file_path)
        print(f"Transactions and summary successfully added to sheet '{sheet_name}' in {file_path}")
    except Exception as e:
        print(f"Error saving file: {e}")